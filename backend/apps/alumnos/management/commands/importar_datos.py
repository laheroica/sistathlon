"""
Management command para importar horarios y alumnos desde los Excel de Athlon.

Uso:
    python manage.py importar_datos

Archivos fuente (rutas hardcodeadas al directorio de Google Drive del proyecto):
    - Control/Profes/HorariosX.xlsx  -> horarios maestros
    - Control/Alumnos/2026 - s107.xlsx  -> alumnos Athlon 107
    - Control/Alumnos/2026- s24.xlsx   -> alumnos Athlon 24
"""

import os
import re
from datetime import datetime, date
from django.core.management.base import BaseCommand
from django.db import transaction

# Ruta base de los archivos fuente
BASE_EXCEL = os.path.expandvars(
    r"G:\Mi unidad\Negocios\Athlon\Control"
)

HORARIOS_FILE = os.path.join(BASE_EXCEL, "Profes", "HorariosX.xlsx")
S107_FILE = os.path.join(BASE_EXCEL, "Alumnos", "2026 - s107.xlsx")
S24_FILE  = os.path.join(BASE_EXCEL, "Alumnos", "2026- s24.xlsx")

DIA_MAP = {
    'lun': 'lun', 'lunes': 'lun',
    'mar': 'mar', 'martes': 'mar',
    'mié': 'mie', 'miér': 'mie', 'miércoles': 'mie', 'miér.': 'mie',
    'jue': 'jue', 'jueves': 'jue',
    'vie': 'vie', 'viernes': 'vie',
    'sáb': 'sab', 'sábado': 'sab',
}

DISC_MAP = {
    'CF': 'CF', 'FC': 'CF',
    'HF': 'HF',
    'HX': 'HX', 'HYROX': 'HX',
    'K':  'KD', 'KIDS': 'KD',
    'TN': 'TN', 'TEENS': 'TN',
}


def excel_serial_to_date(serial):
    """Convierte número serial de Excel a date de Python."""
    if not serial or not isinstance(serial, (int, float)):
        return None
    try:
        return datetime(1899, 12, 30) + __import__('datetime').timedelta(days=int(serial))
    except Exception:
        return None


def normalizar_dis(dis_str):
    """'CF x3' -> {'disciplina': 'CF', 'frecuencia': '3x'}"""
    if not dis_str:
        return {'disciplina': 'CF', 'frecuencia': '3x'}
    d = dis_str.upper().replace(' ', '')
    if 'TEENS' in d:
        return {'disciplina': 'TN', 'frecuencia': '3x'}
    if 'KIDS' in d:
        return {'disciplina': 'KD', 'frecuencia': '3x'}
    if 'HYROX' in d or d.startswith('HX'):
        return {'disciplina': 'HX', 'frecuencia': '3x'}
    disc = 'HF' if d.startswith('HF') else 'CF'
    freq = '3x'
    if 'X2' in d:
        freq = '2x'
    elif 'X5' in d:
        freq = 'libre'  # 5x → Pase Libre
    elif 'X3' in d:
        freq = '3x'
    return {'disciplina': disc, 'frecuencia': freq}


def normalizar_nombre_profe(nombre):
    """'FLOR' -> 'Flor', 'Damian' -> 'Damian'"""
    if not nombre or not isinstance(nombre, str):
        return None
    n = nombre.strip()
    if not n or n == '?':
        return None
    # Mapeo de alias usados en el Excel
    alias = {
        'FLOR': 'Flor',
        'MARIO': 'Mario',
        'BARBI': 'Barbi',
        'DAMIÁN': 'Damián',
        'SUGUS': 'Sugus',
        'SOFI': 'Sofi',
        'MAXI': 'Maxi',
        'BRUNO': 'Bruno',
        'DENI': 'Deni',
        'DAY': 'Day',
    }
    return alias.get(n.upper(), n.title())


class Command(BaseCommand):
    help = 'Importa horarios maestros y alumnos desde los Excel de Athlon'

    def add_arguments(self, parser):
        parser.add_argument('--solo-horarios', action='store_true', help='Solo importar horarios')
        parser.add_argument('--solo-alumnos', action='store_true', help='Solo importar alumnos')
        parser.add_argument('--limpiar', action='store_true', help='Borrar registros existentes antes de importar')

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write('ERROR: openpyxl no está instalado. Ejecutá: pip install openpyxl')
            return

        solo_horarios = options['solo_horarios']
        solo_alumnos = options['solo_alumnos']

        if not solo_alumnos:
            self.importar_horarios(openpyxl, options['limpiar'])

        if not solo_horarios:
            self.importar_alumnos(openpyxl, options['limpiar'])
            self.importar_pagos_abril(openpyxl)

        self.stdout.write(self.style.SUCCESS('\nImportacion completa.'))

    # --- HORARIOS MAESTROS ---
    def importar_horarios(self, openpyxl, limpiar):
        from apps.horarios.models import HorarioMaestro
        from apps.profes.models import Profe

        self.stdout.write('\nImportando horarios maestros...')

        if limpiar:
            HorarioMaestro.objects.all().delete()
            self.stdout.write('  -Registros anteriores eliminados.')

        try:
            wb = openpyxl.load_workbook(HORARIOS_FILE, data_only=True)
        except FileNotFoundError:
            self.stderr.write(f'  ERROR: No se encontró {HORARIOS_FILE}')
            return

        ws = wb['Hoja1']
        rows = list(ws.iter_rows(values_only=True))

        # Parsear las dos grillas (107 y 24) dentro de la hoja
        # Cada grilla tiene: fila título, fila disciplinas, fila horas, filas días
        creados = 0
        actualizados = 0

        def parse_grilla(title_row_idx, sede):
            nonlocal creados, actualizados
            disc_row = rows[title_row_idx + 1]
            hora_row = rows[title_row_idx + 2]

            # Construir mapa col → (disciplina, hora_str)
            turnos = {}
            for col_i, (disc, hora) in enumerate(zip(disc_row, hora_row)):
                if disc and hora and str(disc).strip().upper() in DISC_MAP:
                    hora_int = int(hora) if isinstance(hora, (int, float)) else None
                    if hora_int:
                        hora_str = f"{hora_int:02d}:00"
                        disc_norm = DISC_MAP[str(disc).strip().upper()]
                        turnos[col_i] = (disc_norm, hora_str)

            # Leer filas de días (5 días después del header)
            for day_offset in range(5):
                row = rows[title_row_idx + 3 + day_offset]
                if not row:
                    continue
                dia_raw = str(row[4]).strip().lower() if row[4] else ''
                dia = DIA_MAP.get(dia_raw)
                if not dia:
                    continue

                for col_i, (disc, hora_str) in turnos.items():
                    if col_i >= len(row):
                        continue
                    profe_raw = row[col_i]
                    profe_nombre = normalizar_nombre_profe(str(profe_raw) if profe_raw else None)

                    profe_obj = None
                    if profe_nombre:
                        profe_obj, _ = Profe.objects.get_or_create(
                            nombre=profe_nombre,
                            defaults={
                                'color': '#6b7280',
                                'sede': 'ambas',
                                'tipo_liquidacion': 'hora',
                                'fecha_inicio': date.today(),
                            }
                        )

                    obj, created = HorarioMaestro.objects.update_or_create(
                        sede=sede,
                        dia=dia,
                        hora=hora_str,
                        disciplina=disc,
                        defaults={'profe': profe_obj, 'activo': True}
                    )
                    if created:
                        creados += 1
                    else:
                        actualizados += 1

        # Encontrar filas de título dinámicamente
        for i, row in enumerate(rows):
            if row[4] and 'ATHLON 107' in str(row[4]).upper():
                parse_grilla(i, '107')
                break  # solo tomar el primer bloque

        for i, row in enumerate(rows):
            if row[4] and 'ATHLON 24' in str(row[4]).upper():
                parse_grilla(i, '24')
                break

        self.stdout.write(f'  OK:Horarios: {creados} creados, {actualizados} actualizados.')

    # --- ALUMNOS ---
    def importar_alumnos(self, openpyxl, limpiar):
        from apps.alumnos.models import Alumno, EstadoAlumno

        self.stdout.write('\nImportando alumnos...')

        if limpiar:
            Alumno.objects.all().delete()
            self.stdout.write('  -Alumnos anteriores eliminados.')

        total_creados = 0
        total_skip = 0

        for path, sheet_name, sede in [
            (S107_FILE, '2026- s107', '107'),
            (S24_FILE,  'LISTADO ALUMNOS', '24'),
        ]:
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
            except FileNotFoundError:
                self.stderr.write(f'  ERROR: No se encontró {path}')
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            creados = 0
            skip = 0

            with transaction.atomic():
                for row in rows[5:]:  # saltar encabezados
                    nombre_raw = row[5] if len(row) > 5 else None
                    if not nombre_raw or not isinstance(nombre_raw, str) or len(nombre_raw.strip()) < 3:
                        continue

                    estado_raw = row[2] if len(row) > 2 else None
                    if estado_raw not in ('Vencido', 'Vigente'):
                        continue

                    # Datos de pago para determinar estado
                    pago_abril = row[14] if len(row) > 14 else None
                    pago_mayo  = row[12] if len(row) > 12 else None
                    vto        = row[3]  if len(row) > 3  else None

                    # Determinar estado según antigüedad del último pago
                    # vto es número serial Excel — 46023 = 01/02/2026
                    tiene_pago_mayo  = pago_mayo not in (None, '')
                    tiene_pago_abril = pago_abril not in (None, '')
                    vto_reciente     = isinstance(vto, (int, float)) and vto > 46023  # posterior a feb 2026

                    if tiene_pago_mayo:
                        estado_inicial = 'activo'
                    elif tiene_pago_abril:
                        estado_inicial = 'mora'       # pagó abril, no mayo (~30 días)
                    elif vto_reciente:
                        estado_inicial = 'mora'       # venció hace poco
                    else:
                        estado_inicial = 'alejado'    # sin pagos recientes

                    dis_raw    = row[7] if len(row) > 7 else ''
                    cuota      = float(row[8]) if len(row) > 8 and row[8] else 0
                    horario_h  = row[1] if len(row) > 1 else None
                    fecha_ini_s = row[4] if len(row) > 4 else None

                    nombre = nombre_raw.strip()
                    horario = f"{int(horario_h):02d}:00" if isinstance(horario_h, (int, float)) else ''
                    dis_info = normalizar_dis(str(dis_raw) if dis_raw else '')

                    fecha_inicio = None
                    if isinstance(fecha_ini_s, (int, float)):
                        dt = excel_serial_to_date(fecha_ini_s)
                        fecha_inicio = dt.date() if dt else date(2024, 1, 1)
                    elif isinstance(fecha_ini_s, (datetime, date)):
                        fecha_inicio = fecha_ini_s if isinstance(fecha_ini_s, date) else fecha_ini_s.date()
                    else:
                        fecha_inicio = date(2024, 1, 1)

                    # Dividir nombre / apellido heurísticamente
                    partes = nombre.split(' ')
                    if len(partes) >= 2:
                        primer_nombre = partes[0]
                        apellido = ' '.join(partes[1:])
                    else:
                        primer_nombre = nombre
                        apellido = ''

                    # Usar nombre + sede como clave única (no tenemos DNI)
                    if not Alumno.objects.filter(nombre=primer_nombre, apellido=apellido, sede=sede).exists():
                        import uuid
                        Alumno.objects.create(
                            nombre=primer_nombre,
                            apellido=apellido,
                            dni=f'IMPORT-{sede}-{uuid.uuid4().hex[:8]}',  # placeholder único
                            celular='',
                            sede=sede,
                            fecha_inicio=fecha_inicio,
                            disciplina=dis_info['disciplina'],
                            frecuencia=dis_info['frecuencia'],
                            horario=horario,
                            cuota_actual=cuota,
                            estado=estado_inicial,
                            tipo_precio='regular',
                        )
                        creados += 1
                    else:
                        skip += 1

            self.stdout.write(f'  OK:Sede {sede}: {creados} alumnos creados, {skip} ya existían.')
            total_creados += creados
            total_skip += skip

        self.stdout.write(f'\n  Total: {total_creados} alumnos importados, {total_skip} omitidos.')

    # --- PAGOS DE ABRIL ---
    def importar_pagos_abril(self, openpyxl):
        from apps.alumnos.models import Alumno
        from apps.pagos.models import Pago
        from datetime import date

        self.stdout.write('\nImportando pagos de Abril 2026...')
        mes_abril = date(2026, 4, 1)
        creados = 0
        sin_alumno = 0

        for path, sheet_name, sede in [
            (S107_FILE, '2026- s107', '107'),
            (S24_FILE,  'LISTADO ALUMNOS', '24'),
        ]:
            try:
                wb = openpyxl.load_workbook(path, data_only=True)
            except FileNotFoundError:
                continue

            ws = wb[sheet_name]
            rows = list(ws.iter_rows(values_only=True))

            with transaction.atomic():
                for row in rows[5:]:
                    nombre_raw = row[5] if len(row) > 5 else None
                    if not nombre_raw or not isinstance(nombre_raw, str):
                        continue

                    # col 14 = fecha pago Abril, col 15 = monto Abril
                    fecha_pago_s = row[14] if len(row) > 14 else None
                    monto_abril  = row[15] if len(row) > 15 else None

                    if not fecha_pago_s or not monto_abril:
                        continue

                    try:
                        monto = float(monto_abril)
                        if monto <= 0:
                            continue
                    except (TypeError, ValueError):
                        continue

                    fecha_pago = None
                    if isinstance(fecha_pago_s, (int, float)):
                        dt = excel_serial_to_date(fecha_pago_s)
                        fecha_pago = dt.date() if dt else mes_abril
                    elif isinstance(fecha_pago_s, date):
                        fecha_pago = fecha_pago_s
                    else:
                        fecha_pago = mes_abril

                    nombre = nombre_raw.strip()
                    partes = nombre.split(' ')
                    primer_nombre = partes[0]
                    apellido = ' '.join(partes[1:]) if len(partes) >= 2 else ''

                    alumno = Alumno.objects.filter(
                        nombre=primer_nombre, apellido=apellido, sede=sede
                    ).first()

                    if not alumno:
                        sin_alumno += 1
                        continue

                    if not Pago.objects.filter(alumno=alumno, mes=mes_abril).exists():
                        Pago.objects.create(
                            alumno=alumno,
                            mes=mes_abril,
                            monto=monto,
                            fecha_pago=fecha_pago,
                            metodo='efectivo',
                        )
                        creados += 1

        self.stdout.write(f'  OK: {creados} pagos de Abril importados ({sin_alumno} sin alumno match).')
